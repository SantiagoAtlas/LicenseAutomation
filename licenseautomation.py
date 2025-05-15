import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
import pdfplumber
from collections import defaultdict

# File paths
checklist_file_path = 'C:\\Data\\LicenseAutomation\\System_Checklist_Example.xlsm'
system_list_file_path = 'C:\\Data\\LicenseAutomation\\System_List_Example.xlsx'
pdf_sys_path = 'C:\\Data\\LicenseAutomation\\30028604-2306.pdf'
pdf_enso_path = 'C:\\Data\\LicenseAutomation\\30028003 - 1103.pdf'

# Headers with new PDF data columns added
headers = [
    'SAP Kommissinsnummer', 'SAP Position', 'Material-Nummer vom Schrank', 'Firma', 'Standort', 'Projekt',
    'Steuerung', 
    'V3 – Board Linux Version',
    'V3 – Board Firmware',
    'V3 – Board Seriennummer',
    'V3-Board: Hardware - Version',
    'Verwendeter IPC Typ',
    'Seriennummer IPC 1',
    'Seriennummer V3-Board 1',
    'Komponente', 'Schnittstellen', 'Stationsname', 'IP-Adresse (Kundennetz)',
    'MAC-Adersse 1', 'Lizenznummer (S/N)', 'Lizenz', 'Auslauf-Datum', 'Kommentar', 'Multiplicity'
]

def extract_values_from_pdf(pdf_path, keywords):
    extracted = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    for key in keywords:
                        if row and row[0] and key in row[0]:
                            extracted[key] = row[1] if len(row) > 1 else ''
    return extracted

print("📂 Opening Excel files...")
checklist_workbook = openpyxl.load_workbook(checklist_file_path, keep_vba=True)
main_info_sheet = checklist_workbook['Main Info']
systemmatrix_sheet = checklist_workbook['Systemmatrix']
system_list_workbook = openpyxl.load_workbook(system_list_file_path)
overview_sheet = system_list_workbook['Overview'] if 'Overview' in system_list_workbook.sheetnames else system_list_workbook.create_sheet('Overview')

# Read Main Info
commission_no = main_info_sheet['B3'].value
end_customer = main_info_sheet['B7'].value
location = main_info_sheet['B8'].value
project_name = main_info_sheet['B5'].value
system = main_info_sheet['B11'].value
steuerung = main_info_sheet['B9'].value

print("📋 Retrieved Main Info:")
print(f"   🔹 Commission No: {commission_no}")
print(f"   🔹 Customer: {end_customer}")
print(f"   🔹 Location: {location}")
print(f"   🔹 Project: {project_name}")
print(f"   🔹 System: {system}")
print(f"   🔹 Steuerung: {steuerung}")

# Select PDF and keywords based on System
if system == 'SYS6000':
    selected_pdf = pdf_sys_path
    keywords = [
        'V3 – Board Linux Version',
        'V3 – Board Firmware',
        'V3 – Board Seriennummer',
        'V3-Board: Hardware - Version'
    ]
elif system == 'ENSO7000':
    selected_pdf = pdf_enso_path
    keywords = [
        'Verwendeter IPC Typ',
        'Seriennummer IPC 1',
        'Seriennummer V3-Board 1'
    ]
else:
    selected_pdf = None
    keywords = []
    print("⚠️ Unknown System value. No PDF selected.")

# Extract info from PDF
pdf_info = {}
if selected_pdf:
    print(f"🔍 Extracting data from PDF: {selected_pdf}")
    pdf_info = extract_values_from_pdf(selected_pdf, keywords)
    for key, val in pdf_info.items():
        print(f"   ✅ {key}: {val}")
else:
    print("❌ No PDF extraction performed due to missing or invalid System.")

# Create headers if missing
existing_headers = [cell.value for cell in overview_sheet[1]] if overview_sheet.max_row > 0 else []
if all(h in existing_headers for h in headers):
    print("✅ Headers already exist.")
else:
    if overview_sheet.max_row > 0:
        overview_sheet.delete_rows(1, 1)
    overview_sheet.append(headers)
    center_alignment = Alignment(horizontal='center')
    for col in range(1, len(headers) + 1):
        cell = overview_sheet.cell(row=1, column=col)
        cell.alignment = center_alignment
        cell.font = Font(bold=True)
    print("🆕 Headers created.")

# Check existing commission entries
existing_rows = list(overview_sheet.iter_rows(min_row=2, values_only=True))
existing_positions = {row[1] for row in existing_rows if row and row[0] == commission_no}

# Group columns with identical X marks
print("🔍 Analyzing Systemmatrix for duplications...")
start_col_index = column_index_from_string('G')
col_signature_map = defaultdict(list)

for col in range(start_col_index, systemmatrix_sheet.max_column + 1):
    sap_position = systemmatrix_sheet.cell(row=4, column=col).value
    if not sap_position:
        continue

    signature = tuple(
        row for row in range(5, systemmatrix_sheet.max_row + 1)
        if isinstance(systemmatrix_sheet.cell(row=row, column=col).value, str)
        and systemmatrix_sheet.cell(row=row, column=col).value.strip().lower() == 'x'
    )
    col_signature_map[signature].append((col, sap_position))

# Process groups and add rows
added_rows = 0
skipped_rows = 0
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
right_alignment = Alignment(horizontal='right')

for signature, columns in col_signature_map.items():
    first_col, first_position = columns[0]

    if first_position in existing_positions:
        print(f"⏭️ Position {first_position} already exists for commission {commission_no}, skipping.")
        skipped_rows += 1
        continue

    item_row = signature[0] if signature else None
    item_name = systemmatrix_sheet.cell(row=item_row, column=2).value if item_row else None
    order_article_number = systemmatrix_sheet.cell(row=item_row, column=3).value if item_row else None
    combined_value = f"{item_name} / {order_article_number}" if item_name and order_article_number else item_name or order_article_number

    multiplicity = len(columns)

    # Construct the new row with PDF data mapped to new columns
    new_row = [
        commission_no, first_position, combined_value, end_customer, location,
        project_name, steuerung,
        pdf_info.get('V3 – Board Linux Version', ''),
        pdf_info.get('V3 – Board Firmware', ''),
        pdf_info.get('V3 – Board Seriennummer', ''),
        pdf_info.get('V3-Board: Hardware - Version', ''),
        pdf_info.get('Verwendeter IPC Typ', ''),
        pdf_info.get('Seriennummer IPC 1', ''),
        pdf_info.get('Seriennummer V3-Board 1', ''),
        '', '', '', '', '', '', '', '', '', multiplicity
    ]

    overview_sheet.append(new_row)
    target_row = overview_sheet.max_row
    for col_index in range(1, len(headers) + 1):
        cell = overview_sheet.cell(row=target_row, column=col_index)
        cell.border = thin_border
        cell.alignment = right_alignment

    print(f"➕ Added position {first_position} with multiplicity {multiplicity}.")
    added_rows += 1

print(f"\n🧾 Summary for Commission {commission_no}:")
print(f"   ✅ Rows added: {added_rows}")
print(f"   ⏭️ Positions skipped: {skipped_rows}")

# Add dropdown to Interfaces column ('Schnittstellen'), which is now at position 15 (indexing from 1)
print("🔧 Adding dropdown to 'Interfaces' column...")
data_validation = DataValidation(type="list", formula1='"MQTT,OPC UA,IBM MQ"', showDropDown=True)
overview_sheet.add_data_validation(data_validation)
for row in range(2, overview_sheet.max_row + 1):
    cell = overview_sheet[f'O{row}']  # 'Schnittstellen' is column 15 = 'O'
    data_validation.add(cell)
    cell.border = thin_border
    cell.alignment = right_alignment

# Adjust column widths
print("📐 Adjusting column widths...")
for col in range(1, len(headers) + 1):
    max_length = max((len(str(cell.value)) if cell.value else 0) for cell in overview_sheet[get_column_letter(col)])
    overview_sheet.column_dimensions[get_column_letter(col)].width = max_length + 2

overview_sheet.auto_filter.ref = overview_sheet.dimensions

# Save file
try:
    system_list_workbook.save(system_list_file_path)
    print("💾 File saved successfully. Process completed ✅")
except PermissionError:
    print("❌ Error: Could not save file. Please close the Excel file if it's open.")
